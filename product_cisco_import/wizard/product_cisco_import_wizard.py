# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Addons modules by CLEARCORP S.A.
#    Copyright (C) 2009-TODAY CLEARCORP S.A. (<http://clearcorp.co.cr>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

from openerp.osv import osv, fields
import base64
from openerp.tools.translate import _
from openpyxl import load_workbook, Workbook
import StringIO

STATES = [('update','Update'),('new','New')]

class ProductCiscoImportLine(osv.TransientModel):

    _name = 'product.cisco.import.line'

    def _get_state(self, cr, uid, ids, field_name, arg, context=None):
        product_obj = self.pool.get('product.product')
        res = {}
        for line in self.browse(cr ,uid,ids, context=context):
            product_id = product_obj.search(cr, uid, [('default_code','=',line.default_code)])
            if product_id:
                res[line.id] = {
                                'product_state':'update',
                                'product_id': product_id,
                                }
            else:
                res[line.id] = {
                                'product_state':'new',
                                'product_id': False,
                                }
        return res

    _columns = {
        'name': fields.char('Product Name'),
        'default_code': fields.char('Product Code'),
        'standard_price': fields.float('Price in US Dolar'),
        'product_state': fields.function(_get_state, selection=STATES, string='Product Status', type='selection', readonly=True, store=True, multi='products_line'),
        'wizard_id': fields.many2one('product.cisco.import.wizard','wizard'),
        'product_id': fields.function(_get_state, string='Products', readonly=True, type='many2one', obj='product.product', store=True, multi='products_line'),
        }

class ProductCiscoImportWizard(osv.TransientModel):

    _name = 'product.cisco.import.wizard'

    _columns = {
                'file': fields.binary('Products File'),
                'state': fields.selection([ ('draft', 'Draft'), ('validate', 'Validate')], 'State', readonly=True),
                'line_ids': fields.one2many('product.cisco.import.line', 'wizard_id', string='Products Cisco'),
    } 

    _defaults = {
        'state': 'draft'
    }

    def create_product(self, cr, uid, ids, context=None):
        wizard = self.browse(cr, uid, ids[0], context=context)
        product_obj = self.pool.get('product.product')
        product_tmpl_obj = self.pool.get('product.template')
        products = []
        for line in wizard.line_ids:
            values={
                'default_code': line.default_code,
                'name': line.name,
                'standard_price': line.standard_price,
                }
            if line.product_id:
            #if product_id:
                product_obj.write(cr, uid, line.product_id.id, values, context=context)
                products.append(line.product_id.id)
            else:
                products_new = product_obj.create(cr, uid, values, context=context)
                products.append(products_new)
        
        products_ids = product_obj.search(cr, uid, [('id','not in',products)])
        product_obj.write(cr, uid, products_ids, {'active': False}, context=context)
        
        product_tmpl_ids = product_tmpl_obj.search(cr, uid, [('product_variant_ids','in',products_ids)])
        product_tmpl_obj.write(cr, uid, product_tmpl_ids, {'active': False}, context=context)
        mod_obj = self.pool.get('ir.model.data')
        res = mod_obj.get_object_reference(cr, uid, 'product', 'product_product_tree_view')
        return {
           'name': _('Imported products'),
           'type': 'ir.actions.act_window',
           'res_model': 'product.product',
           'view_type': 'form',
           'view_mode': 'tree',
           'view_id': [res and res[1] or False],
           'context': context,
           'domain': [('id','in',products)],
            }

    def import_data(self,cr,uid,ids,context=None):
        wizard = self.browse(cr, uid, ids[0], context=context)
        lines_obj=self.pool.get('product.cisco.import.line')
         # Decode the base64 encoded file
        data = base64.decodestring(wizard.file)
        # Gather IO from the decoded string
        file = StringIO.StringIO()
        file.write(data)
        # Load workbook from IO
        workbook = load_workbook(file)
        worksheets = workbook.get_sheet_names()
        # Get the name of sheets in document
        for sheetname in worksheets:
            worksheet = workbook.get_sheet_by_name(sheetname)
            data_rows = worksheet.rows[6:]
            #Data begin in this cell, skip header
            row_number = 6
            for row in data_rows:
                if worksheet.cell(row=row_number, column=5).value == None:
                    row_number+=1
                    continue
                str = worksheet.cell(row=row_number, column=5).value
                str = str.replace(',','')
                standard_str = str.replace("'",'')
                standard_str = str.replace('$','')
                standard_price = standard_str.replace(' ','')
                standard_price = float(standard_price)
                vals = {
                    'default_code' : worksheet.cell(row=row_number, column=2).value,
                    'name' : worksheet.cell(row=row_number, column=3).value,
                    'standard_price' : standard_price,
                    'wizard_id': wizard.id,
                 }
                lines_obj.create(cr,uid,vals,context=context)
                row_number+=1
        wizard.write({'state':'validate'})
        return {
           'name': _('Import Lines'),
           'type': 'ir.actions.act_window',
           'res_model': self._name,
           'view_type': 'form',
           'view_mode': 'form',
           'target': 'new',
           'context': context,
           'res_id': wizard.id,
            }
